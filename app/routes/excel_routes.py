from fastapi import APIRouter, UploadFile, HTTPException, status, Depends
from openpyxl import load_workbook
from app.database import excel_collection, user_collection
from app.models.excel_model import ExcelSheetModel
from app.auth import oauth2_scheme, get_user
from pydantic import ValidationError
import hashlib
import logging

# Inicializar el logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

router = APIRouter()

# Función para calcular el hash del archivo Excel
def calculate_file_hash(file: UploadFile):
    hasher = hashlib.sha256()
    buffer = file.file.read()
    hasher.update(buffer)
    file.file.seek(0)  # Resetear la posición del archivo para ser leído nuevamente
    return hasher.hexdigest()

@router.post("/upload-excel", status_code=201)
async def upload_excel(file: UploadFile, token: str = Depends(oauth2_scheme)):
    # Verificar que el tipo de archivo sea Excel
    if file.content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        logger.error("Invalid file type.")
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel files are allowed.")

    # Obtener el usuario a partir del token
    user = await get_user(token)  
    if user is None:
        raise HTTPException(status_code=403, detail="Could not validate credentials.")

    try:
        # Calcular el hash del archivo para verificar idempotencia
        file_hash = calculate_file_hash(file)

        # Verificar si ya existe un archivo con el mismo hash en la base de datos
        existing_file = await excel_collection.find_one({"file_hash": file_hash})
        if existing_file:
            logger.info("File already uploaded.")
            raise HTTPException(status_code=400, detail="This Excel file has already been uploaded.")

        # Leer el archivo de Excel
        workbook = load_workbook(file.file)
        sheet = workbook.active

        # Obtener los datos de la hoja
        columns = [cell.value for cell in sheet[1]]  # Cabecera (primera fila)
        rows = [
            {columns[i]: cell.value for i, cell in enumerate(row)} 
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ]

        # Validar que haya datos en la hoja
        if not columns or not rows:
            logger.error("Excel file is empty or malformed.")
            raise HTTPException(status_code=400, detail="Excel file is empty or malformed.")

        # Validar si las columnas necesarias están presentes
        required_columns = {"Name", "Age", "City"}
        if not required_columns.issubset(set(columns)):
            missing_columns = required_columns - set(columns)
            logger.error(f"Missing required columns: {missing_columns}")
            raise HTTPException(status_code=400, detail=f"Missing required columns: {missing_columns}")

        # Crear el modelo de la hoja de Excel
        excel_data = ExcelSheetModel(sheet_name=sheet.title, columns=columns, rows=rows)

        # Guardar los datos en MongoDB junto con el hash del archivo
        result = await excel_collection.insert_one({
            **excel_data.dict(),
            "file_hash": file_hash  # Guardar el hash del archivo para validación futura
        })

        logger.info(f"Excel sheet '{sheet.title}' uploaded successfully with ID {result.inserted_id}.")
        return {"message": "Excel sheet uploaded successfully", "id": str(result.inserted_id)}

    except ValidationError as ve:
        logger.error(f"Validation error: {ve}")
        raise HTTPException(status_code=400, detail=f"Data validation error: {ve}")
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred: {e}")

@router.get("/protected-route")
async def protected_route(token: str = Depends(oauth2_scheme)):
    return {"message": "You have access to the protected route with a valid token."}
