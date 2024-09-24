from openpyxl import load_workbook
from io import BytesIO

def process_excel(file_content: bytes):
    # Cargar el archivo de Excel
    workbook = load_workbook(filename=BytesIO(file_content))
    sheet = workbook.active

    # Extraer los valores de las celdas (en lugar de los objetos Cell)
    columns = [cell.value for cell in sheet[1]]  # Cabecera (primera fila)
    rows = [
        {columns[i]: cell for i, cell in enumerate(row)}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

    return {
        "sheet_name": sheet.title,
        "columns": columns,
        "rows": rows
    }
